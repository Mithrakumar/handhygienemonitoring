#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Apr  5 22:28:23 2021

@author: Mithra
"""


from openpyxl import load_workbook
workbook = load_workbook(filename="Database_students_VIT.xlsx")
spreadsheet = workbook.active
#print(workbook.sheetnames)
#print(spreadsheet["B2"].value)

from openpyxl import Workbook
import pandas as pd
from openpyxl.utils.dataframe import  dataframe_to_rows  
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import xlrd  
import sys
def quit():
    sys.exit()
    

myPath = r'/Users/Mithra/Desktop/Hand Hygiene Monitoring/Database_students_VIT.xlsx'
for sh in xlrd.open_workbook(myPath).sheets():  
    for row in range(sh.nrows):
        for col in range(sh.ncols):
            myCell = sh.cell(row, col)
            print(myCell)
            if myCell.value == '17BEC0037':
                print('-----------')
                print('Found!')
                password_cell = (xl_rowcol_to_cell(row,col+1))
                print(spreadsheet[password_cell].value)
                check = str(spreadsheet[password_cell].value)
                x = str(input("enter password"))
                if x == check:
                    print("authentication complete")
                else:
                    print("error")
                quit()